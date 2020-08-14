from skimage import io
import matplotlib.pyplot as P
import os
import openpyxl
import numpy as np
from skimage.feature import hog
from skimage.transform import resize
from skimage import data, color, exposure
import glob
from sklearn import svm

def readImage_AND_getHOG(path):
    os.chdir(path)
    imageList=glob.glob("*.jpg")
    print("Number of Images "+str(len(imageList)))
    #Sorting Image
    imageList=sorted(imageList,key=lambda x: int(os.path.splitext(x)[0]))
    print("Founded Image")
    print(imageList)
    hogArray=[]
    i=0
    for image in imageList:
        readed_image = io.imread(image)
        #Resize Image
        readed_image = resize(readed_image, (32,32, 3))
        #Convert RGB to Gray Scale
        gray_image = color.rgb2gray(readed_image)
        # Hog feature
        fd, hog_image = hog(gray_image, orientations=8, pixels_per_cell=(16, 16),
                            cells_per_block=(1, 1), visualise=True)
        hogArray.append(fd)


    hogArray=np.array(hogArray)

    return hogArray


def read_EXCEL_file(path):
    markList=[]
    label= openpyxl.load_workbook(path)
    sheet = label.get_sheet_by_name('Sheet1')
    row=sheet.max_row;
    column=sheet.max_column
    ### Here need to Change, Now Value is set Manually,
    # Excel Sheet Column 1 (picture title) show invalid result
    title=[]
    Image_No = 401
    for i in range(2, row + 1):

        IMAGE_NO1 = sheet.cell(row=i, column=1).value
        #Read Mark from Exel Sheet
        MARK = sheet.cell(row=i, column=2).value
        #print("Name: " + str(IMAGE_NO1) + "  Mark ", str(sheet.cell(row=i, column=2).value))
        markList.append(MARK)
        title.append(Image_No)
        Image_No = Image_No + 1

    markList=np.array(markList)
    title=np.array(title)
    return markList,title



def accuracy_test(image_title, expected_mark, predicted_mark):
    flag=[]
    matched_counter=0
    notMatched_counter=0
    #Holds the distance between Actual mark and predicted mark
    distance=[]
    actual_mark_size=len(expected_mark)
    predicted_mark_size=len(predicted_mark)
    if(actual_mark_size!=predicted_mark_size):
        print("Different size array. Size must be the same")
    elif(actual_mark_size==predicted_mark_size):
        for i in range(actual_mark_size):
            #if predicted mark is matched with actual mark
            if(expected_mark[i]==predicted_mark[i]):
                flag.append(True)
                #Does not have any distance
                print(str(image_title[i])+".tif Matched")
                distance.append(0)
            elif(expected_mark[i]!=predicted_mark[i]):
                flag.append(False)
                diff=abs(expected_mark[i] - predicted_mark[i])
                distance.append(diff.tolist())
                print(str(image_title[i]) +".tif Not Matched Distance " +" Distance found " + str(distance[i]) +" Expected: " + str(expected_mark[i]) + " Predict " + str(predict[i]))

        for i in range(actual_mark_size):
            if(flag[i]==True):
                matched_counter=matched_counter+1
            elif(flag[i]==False):
                notMatched_counter= notMatched_counter+1

        accuracy_in_percentant= ((matched_counter*100)/actual_mark_size)
        return accuracy_in_percentant






### Path of Training and Testing Data ###
dir_path = os.path.dirname(os.path.realpath(__file__))
training_image_path= "path"
testing_image_path="path"
mark_path="path"


training_hog_feature= readImage_AND_getHOG(training_image_path)
testing_hog_feature= readImage_AND_getHOG(testing_image_path)
mark,title=read_EXCEL_file(mark_path)
X=len(training_hog_feature)
Y=len(mark)


#Fit in SVM
classifier = svm.SVC(gamma='auto', C=100)
classifier.fit(training_hog_feature, mark)

predict=[]
total_size= len(training_hog_feature)
#### Prediction for all image ####
for i in range(total_size):
    feature_value=classifier.predict(training_hog_feature[i])
    feature_value=feature_value.tolist()
    predict.append(feature_value)

accuracy= accuracy_test(title.tolist(),mark,predict)
print("Accuracy: "+str(accuracy)+" %")
